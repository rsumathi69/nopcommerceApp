import openpyxl
from openpyxl.styles import Font,PatternFill

def getRowCount(file,sheetName):
    wb=openpyxl.load_workbook(file)
    ws=wb[sheetName]
    return(ws.max_row)

def getColCount(file,sheetName):
    wb=openpyxl.load_workbook(file)
    ws=wb[sheetName]
    return(ws.max_column)

def readData(file,sheetName,rowNum,colNum):
    wb=openpyxl.load_workbook(file)
    ws=wb[sheetName]
    return ws.cell(rowNum,colNum).value


def writeData(file, sheetName, rowNum, colNum,data):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    ws.cell(rowNum,colNum).value=data
    wb.save(file)

def fillGreen(file, sheetName, rowNum, colNum):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    ws.cell(rowNum,colNum).fill=PatternFill(start_color='60b212',
                end_color='60b212',
                fill_type='solid')
    wb.save(file)
    
def fillRed(file, sheetName, rowNum, colNum):
    wb = openpyxl.load_workbook(file)
    ws = wb[sheetName]
    ws.cell(rowNum,colNum).fill=PatternFill(start_color='ff0000',
                end_color='ff0000',
                fill_type='solid')
    wb.save(file)